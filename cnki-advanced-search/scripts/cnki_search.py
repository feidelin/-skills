#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
知网（CNKI）高级检索自动化脚本 - Playwright 版

用法:
  python3 cnki_search.py --keywords "数字化转型 + 数字化变革" --keywords "企业绩效 + 组织绩效"
  python3 cnki_search.py --keywords "平台经济 + 数字平台" --max-results 50 --port 9222
"""

from __future__ import annotations

import argparse
import asyncio
import re
import sys
from datetime import datetime
from pathlib import Path

from playwright.async_api import async_playwright, Page, BrowserContext

OUTPUT_DIR = Path.home() / "Downloads"


# ── 自定义异常 ────────────────────────────────────────────

class CNKIError(Exception):
    pass

class CNKILoginRequired(CNKIError):
    pass

class CNKICaptchaDetected(CNKIError):
    pass

class CNKIExportFailed(CNKIError):
    pass

class CNKINoResults(CNKIError):
    pass

class CNKIParseError(CNKIError):
    pass


# ── CNKISearcher 主类 ─────────────────────────────────────

class CNKISearcher:
    def __init__(self, page: Page, delay_ms: int = 1000, batch_sleep_s: int = 0):
        self.page = page
        self.delay = delay_ms / 1000  # 转为秒
        self.batch_sleep = batch_sleep_s  # 页间冷却秒数

    async def run(self, keyword_groups: list[str], max_results: int = 100,
                  min_groups: int = 2) -> list[dict]:
        print(f"\n[1/10] 导航到知网高级检索页面...")
        await self._navigate_to_advanced_search()

        print(f"[2/10] 检查验证码...")
        await self._check_captcha_and_wait()

        print(f"[3/10] 检查登录状态...")
        await self._ensure_logged_in()

        print(f"[4/10] 选择学术期刊类别...")
        await self._select_journal_category()

        print(f"[5/10] 勾选CSSCI来源...")
        await self._check_cssci_filter()

        # ── 倒剥洋葱：从全组降级，但不低于 min_groups ────────
        groups_to_try = list(keyword_groups)
        combined_articles: list[dict] = []
        seen_titles: set[str] = set()

        async def _export_current_layer(n: int) -> list[dict]:
            """对当前检索结果排序、导出，返回解析后的文章列表。"""
            print(f"[8/10] 按被引量排序...")
            await self._sort_by_citations()
            # 不切换每页50条：保持默认20条，与导出页面每次最多20条的限制对齐
            print(f"[9/10] 逐页选中并导出（最多{n}篇，每页20条）...")
            return await self._select_and_export_by_page(n)

        def _merge(new_articles: list[dict]):
            """去重合并到 combined_articles（按标题去重）。"""
            added = 0
            for a in new_articles:
                key = a.get('title', '').strip()
                if key and key not in seen_titles:
                    seen_titles.add(key)
                    combined_articles.append(a)
                    added += 1
            return added

        while len(groups_to_try) >= 1:
            layer = len(keyword_groups) - len(groups_to_try) + 1
            print(f"\n[6/10] 输入检索关键词（第{layer}层，{len(groups_to_try)}组 AND）: {groups_to_try}...")
            await self._input_keyword_groups(groups_to_try)

            print(f"[7/10] 执行检索...")
            count = await self._execute_search()
            print(f"       找到 {count} 篇论文")

            is_final = (count >= 20) or (len(groups_to_try) <= min_groups)

            if count > 0:
                remaining = max_results - len(combined_articles)
                if remaining > 0:
                    layer_articles = await _export_current_layer(min(count, remaining))
                    added = _merge(layer_articles)
                    print(f"       本层新增 {added} 篇（去重后累计 {len(combined_articles)} 篇）")

            if is_final:
                if count == 0 and len(groups_to_try) <= min_groups:
                    if len(combined_articles) == 0:
                        print(f"\n[!] 已到最少保留组数({min_groups}组)，结果仍为0篇。")
                        print(f"[!] 请检查关键词是否过于冷僻，考虑替换为上位概念后重新检索。")
                        raise CNKINoResults(
                            f"联合检索（{min_groups}组 AND）结果为0，建议扩展关键词至上位概念后重试。"
                        )
                    else:
                        print(f"\n[!] 最底层结果为0篇，但已从上层汇总 {len(combined_articles)} 篇。")
                elif count > 0 and count < 20 and len(groups_to_try) <= min_groups:
                    print(f"\n[!] 已到最少保留组数({min_groups}组)，当前层仅{count}篇（< 20）。")
                    print(f"[!] 建议：扩展当前关键词至上位概念后重新检索，以获取更多相关文献。")
                break

            # 可以继续降组，但先把本层结果已导出，再重新导航
            dropped = groups_to_try.pop()
            print(f"       自动降至第{layer+1}层（去掉组：{dropped[:40]}）")
            await self._navigate_to_advanced_search()
            await self._check_captcha_and_wait()
            await self._select_journal_category()
            await self._check_cssci_filter()
        # ────────────────────────────────────────────────────

        return combined_articles

    async def _navigate_to_advanced_search(self):
        # 无论当前在哪，先跳到 www.cnki.net 确保完整重载（避免 SPA 同 URL no-op）
        try:
            await self.page.goto("https://www.cnki.net", wait_until="domcontentloaded", timeout=20000)
        except Exception:
            await asyncio.sleep(2)

        # 关闭可能弹出的登录框
        for sel in ['[class*="close"]', '.modal-close', '.login-close', '[aria-label="close"]']:
            try:
                el = await self.page.query_selector(sel)
                if el and await el.is_visible():
                    await el.click()
                    await asyncio.sleep(0.5)
                    break
            except Exception:
                pass

        # 用 JS location.href 跳转到高级检索（绕过 Playwright 对 kns.cnki.net 证书校验）
        await self.page.evaluate("() => { location.href = 'https://kns.cnki.net/kns8s/AdvSearch'; }")

        # 等待表单区真正加载完毕——等 .doctype-menus visible（资源类型选项卡出现代表页面就绪）
        try:
            await self.page.wait_for_selector(
                '.doctype-menus',
                state='visible', timeout=20000
            )
        except Exception:
            # fallback：等任意表单元素 attached，再加额外 sleep
            try:
                await self.page.wait_for_selector(
                    '#gradetxt, .grade-search-content, #sen_1_value1',
                    state='attached', timeout=10000
                )
            except Exception:
                pass
            await asyncio.sleep(5)  # 最终兜底

        # 如果被重定向到验证码页，自动轮询等待
        if "/verify/" in self.page.url or "captcha" in self.page.url.lower():
            print("\n[!] 知网要求滑块验证，请在 Chrome 窗口中完成验证（自动等待最多 120 秒）...")
            for _ in range(60):
                await asyncio.sleep(2)
                if "/verify/" not in self.page.url and "captcha" not in self.page.url.lower():
                    break
            await self.page.evaluate("() => { location.href = 'https://kns.cnki.net/kns8s/AdvSearch'; }")
            try:
                await self.page.wait_for_selector('.doctype-menus', state='visible', timeout=20000)
            except Exception:
                try:
                    await self.page.wait_for_selector(
                        '#gradetxt, .grade-search-content, #sen_1_value1',
                        state='attached', timeout=10000
                    )
                except Exception:
                    pass
                await asyncio.sleep(5)

        await asyncio.sleep(self.delay)

        # 强制显示多条件检索区（新版 CNKI 默认隐藏 .grade-search-content）
        await self._show_advanced_form()

    async def _show_advanced_form(self):
        """强制显示 .grade-search-content（多条件行），隐藏单行 .search-condition"""
        await self.page.evaluate("""
            () => {
                const grade = document.querySelector('.grade-search-content');
                const simple = document.querySelector('.search-condition');
                if (grade && window.getComputedStyle(grade).display === 'none') {
                    grade.style.display = 'block';
                }
                if (simple && window.getComputedStyle(simple).display !== 'none') {
                    simple.style.display = 'none';
                }
            }
        """)
        await asyncio.sleep(0.5)

    async def _check_captcha_and_wait(self):
        """使用 getComputedStyle 检测验证码是否真实可见（避免隐藏DOM误判）"""
        captcha_visible = await self.page.evaluate("""
            () => {
                const selectors = ['.tc-widget-wrapper', '.nc-lang-cnt', '#slideBg', '.geetest_widget'];
                for (const sel of selectors) {
                    const el = document.querySelector(sel);
                    if (!el) continue;
                    const style = window.getComputedStyle(el);
                    if (style.display !== 'none' && style.visibility !== 'hidden' && el.offsetHeight > 0) {
                        return true;
                    }
                }
                return false;
            }
        """)
        if captcha_visible:
            print("\n[!] 检测到验证码，请在浏览器中完成滑块验证（自动等待最多 120 秒）...")
            for _ in range(60):
                await asyncio.sleep(2)
                still_visible = await self.page.evaluate("""
                    () => {
                        const selectors = ['.tc-widget-wrapper', '.nc-lang-cnt', '#slideBg', '.geetest_widget'];
                        for (const sel of selectors) {
                            const el = document.querySelector(sel);
                            if (!el) continue;
                            const style = window.getComputedStyle(el);
                            if (style.display !== 'none' && style.visibility !== 'hidden' && el.offsetHeight > 0)
                                return true;
                        }
                        return false;
                    }
                """)
                if not still_visible:
                    print("  [✓] 验证码已通过")
                    break

    async def _ensure_logged_in(self):
        """检查知网登录状态，未登录时自动轮询等待"""
        try:
            logged_in = await self.page.evaluate("() => typeof islogin === 'function' ? islogin() : true")
            if not logged_in:
                raise CNKILoginRequired()
        except CNKILoginRequired:
            print("\n[!] 知网导出功能需要登录，请在浏览器中登录知网账号（自动等待最多 120 秒）...")
            for _ in range(60):
                await asyncio.sleep(2)
                still_not = await self.page.evaluate(
                    "() => typeof islogin === 'function' ? !islogin() : false"
                )
                if not still_not:
                    print("  [✓] 登录成功")
                    break
            else:
                print("[!] 登录等待超时，将继续执行（可能导致导出失败）")
        except Exception:
            pass  # islogin 不存在时忽略

    async def _select_journal_category(self):
        """点击底部 .doctype-menus 中的'学术期刊'选项卡，等待 CSSCI 过滤区渲染完成"""
        await asyncio.sleep(self.delay)
        # 先等待 .doctype-menus 出现，确保表单已渲染
        try:
            await self.page.wait_for_selector('.doctype-menus', state='visible', timeout=15000)
        except Exception:
            print("  [!] .doctype-menus 未出现，尝试继续...")
        try:
            clicked = await self.page.evaluate("""
                () => {
                    // 精确定位 .doctype-menus 中 resource="JOURNAL" 的链接
                    const journalLink = document.querySelector('.doctype-menus a[resource="JOURNAL"]');
                    if (journalLink) { journalLink.click(); return 'doctype-menus:JOURNAL'; }
                    // 备用：.doctype-menus 中文字为"学术期刊"的 a
                    const menuLinks = [...document.querySelectorAll('.doctype-menus a')];
                    const byText = menuLinks.find(a => a.textContent.trim() === '学术期刊');
                    if (byText) { byText.click(); return 'doctype-menus:text'; }
                    return false;
                }
            """)
            print(f"  [i] 学术期刊点击: {clicked}")

            # 等待 CSSCI checkbox（key="CSI"）渲染出来
            try:
                await self.page.wait_for_function(
                    'function(){ return !!document.querySelector(\'input[key="CSI"]\'); }',
                    timeout=12000
                )
                print("  [i] CSSCI checkbox 已就绪")
            except Exception:
                await asyncio.sleep(3)
                print("  [!] 等待 CSSCI checkbox 超时，继续执行")

        except Exception as e:
            print(f"  [!] 选择学术期刊时出错: {e}，继续执行...")

    async def _check_cssci_filter(self):
        """勾选 CSSCI 来源类别：先取消"全部期刊"，再勾选 CSSCI。"""
        await asyncio.sleep(self.delay)
        # 等待 CSSCI checkbox 出现（学术期刊点击后才会渲染）
        try:
            await self.page.wait_for_function(
                'function(){ return !!document.querySelector(\'input[key="CSI"]\'); }',
                timeout=12000
            )
        except Exception:
            await asyncio.sleep(2)
        try:
            result = await self.page.evaluate("""
                () => {
                    const cssiCb = document.querySelector('input[key="CSI"]');
                    if (!cssiCb) return 'not available';

                    // 取消"全部期刊"（name="all"）
                    const allCb = document.querySelector('input[name="all"]');
                    if (allCb && allCb.checked) {
                        allCb.click();
                    }

                    // 勾选 CSSCI
                    if (!cssiCb.checked) {
                        cssiCb.click();
                    }
                    return cssiCb.checked ? 'checked:CSI' : 'click-sent:CSI';
                }
            """)
            print(f"  [i] CSSCI勾选: {result}")
            await asyncio.sleep(self.delay)
        except Exception as e:
            print(f"  [!] 勾选CSSCI时出错: {e}，继续执行...")

    async def _input_keyword_groups(self, groups: list[str]):
        """填入关键词：兼容新版 sen_N_value1 结构和旧版 #gradetxt 结构"""
        await asyncio.sleep(self.delay)

        # #sen_N_value1 存在但是隐藏的后备字段；真正可见的输入框在 #gradetxt dd .input-box input
        # 只有 #sen_1_value1 可见时才走新路径
        use_new = await self.page.evaluate("""
            () => {
                const el = document.querySelector('#sen_1_value1');
                return !!(el && el.offsetParent !== null);
            }
        """)

        if use_new:
            # ── 新版结构（sen_N_value1）──
            # 用 JS 先解除隐藏/disabled，再用 Playwright 键盘输入（确保 Vue v-model 捕获）
            async def fill_field(idx: int, val: str) -> str:
                sel = f'#sen_{idx}_value1'
                # 解除遮蔽
                await self.page.evaluate(f"""
                    () => {{
                        const el = document.querySelector('{sel}');
                        if (!el) return;
                        el.removeAttribute('disabled');
                        el.removeAttribute('readonly');
                        el.style.display = '';
                        el.style.visibility = 'visible';
                        el.style.opacity = '1';
                        el.style.pointerEvents = 'auto';
                    }}
                """)
                await asyncio.sleep(0.3)
                try:
                    # Playwright 原生 click + fill（会触发 Vue input 事件）
                    await self.page.click(sel, timeout=5000)
                    await asyncio.sleep(0.2)
                    await self.page.keyboard.press("Meta+a")
                    await self.page.keyboard.type(val, delay=30)
                    await asyncio.sleep(0.3)
                    actual = await self.page.evaluate(f"() => document.querySelector('{sel}')?.value || ''")
                    return f'keyboard-filled: {actual[:30]}'
                except Exception as e:
                    # fallback: JS native setter
                    r = await self.page.evaluate(f"""
                        () => {{
                            const el = document.querySelector('{sel}');
                            if (!el) return 'not found';
                            const nativeSetter = Object.getOwnPropertyDescriptor(
                                window.HTMLInputElement.prototype, 'value'
                            ).set;
                            nativeSetter.call(el, {repr(val)});
                            el.dispatchEvent(new InputEvent('input', {{bubbles:true, data:{repr(val)}}}));
                            el.dispatchEvent(new Event('change', {{bubbles:true}}));
                            return 'js-filled: ' + el.value.substring(0, 30);
                        }}
                    """)
                    return r

            r = await fill_field(1, groups[0])
            print(f"  [i] 第1组填入: {r}")
            await asyncio.sleep(self.delay)

            for i, group in enumerate(groups[1:], 2):
                # 点击"+"新增一行
                clicked = await self.page.evaluate("""
                    () => {
                        const btns = [...document.querySelectorAll('a,button,span,i')];
                        const btn = btns.find(b =>
                            b.textContent.trim() === '+' ||
                            (b.className && b.className.includes('add')));
                        if (btn) { btn.click(); return true; }
                        return false;
                    }
                """)
                if not clicked:
                    print(f"  [!] 第{i}组：未找到+按钮，跳过")
                    break
                await asyncio.sleep(self.delay)
                r = await fill_field(i, group)
                print(f"  [i] 第{i}组填入: {r}")
                await asyncio.sleep(self.delay)

        else:
            # ── 旧版结构（#gradetxt）—— 可见的 input 在 #gradetxt dd .input-box input ──
            first_sel = '#gradetxt dd .input-box input[type="text"], #gradetxt dd input[type="text"]'
            try:
                await self.page.click(first_sel, timeout=5000)
                await self.page.keyboard.press("Meta+a")
                await self.page.keyboard.type(groups[0], delay=30)
                actual = await self.page.evaluate(
                    f"() => document.querySelector('{first_sel}')?.value || ''"
                )
                print(f"  [i] 第1组填入(keyboard): {actual[:30]}")
            except Exception as e:
                result0 = await self.page.evaluate(f"""
                    () => {{
                        const sel = '#gradetxt dd .input-box input[type="text"], #gradetxt dd input[type="text"]';
                        const input = document.querySelector(sel);
                        if (!input) return 'not found';
                        input.value = {repr(groups[0])};
                        input.dispatchEvent(new Event('input', {{bubbles: true}}));
                        input.dispatchEvent(new Event('change', {{bubbles: true}}));
                        return 'js-filled: ' + input.value.substring(0, 30);
                    }}
                """)
                print(f"  [i] 第1组填入: {result0}")
            await asyncio.sleep(self.delay)

            for i, group in enumerate(groups[1:], 2):
                clicked = await self.page.evaluate("""
                    () => {
                        const btn = document.querySelector('#gradetxt a.add-group');
                        if (btn) { btn.click(); return true; }
                        return false;
                    }
                """)
                if not clicked:
                    print(f"  [!] 第{i}组：未找到 #gradetxt a.add-group，跳过")
                    break
                await asyncio.sleep(self.delay)
                result = await self.page.evaluate(f"""
                    () => {{
                        const dds = document.querySelectorAll('#gradetxt dd');
                        const lastDd = dds[dds.length - 1];
                        if (!lastDd) return 'no dd found';
                        const input = lastDd.querySelector('input[type="text"]');
                        if (!input) return 'no input in last dd';
                        input.value = {repr(group)};
                        input.dispatchEvent(new Event('input', {{bubbles: true}}));
                        input.dispatchEvent(new Event('change', {{bubbles: true}}));
                        return 'filled: ' + input.value.substring(0, 30);
                    }}
                """)
                print(f"  [i] 第{i}组填入: {result}")
                await asyncio.sleep(self.delay)

    async def _execute_search(self) -> int:
        """点击检索按钮并等待结果，返回结果数量"""
        try:
            search_btn = self.page.locator('button:has-text("检索"), input[value="检索"], .btn-search').first
            await search_btn.click(timeout=5000)
        except Exception:
            await self.page.evaluate("""
                () => {
                    const btns = [...document.querySelectorAll('button, input[type="submit"], a')];
                    const btn = btns.find(b => b.textContent.trim() === '检索' || b.value === '检索');
                    if (btn) btn.click();
                }
            """)

        # 等待结果页加载
        try:
            await self.page.wait_for_url(re.compile(r"kns.cnki.net.*(result|search)"), timeout=20000)
        except Exception:
            await asyncio.sleep(3)

        # 等待结果计数元素出现（SPA 异步渲染，需要真正等待而非固定 sleep）
        try:
            await self.page.wait_for_function("""
                () => {
                    const el = document.querySelector('.pager .count, #countPageDiv, .result-count, .total-count');
                    if (el && el.textContent.trim()) return true;
                    const all = [...document.querySelectorAll('*')];
                    return all.some(e => e.children.length === 0 &&
                        e.textContent.includes('共') && e.textContent.includes('条结果'));
                }
            """, timeout=15000)
        except Exception:
            await asyncio.sleep(self.delay * 2)

        await self._check_captcha_and_wait()

        # 获取结果数量
        try:
            debug_info = await self.page.evaluate("""
                () => {
                    const url = location.href;
                    const el = document.querySelector('.pager .count, #countPageDiv, .result-count, .total-count');
                    const elText = el ? el.textContent.trim() : 'none';
                    const all = [...document.querySelectorAll('*')];
                    const el2 = all.find(e => e.children.length === 0 &&
                        e.textContent.includes('共') && e.textContent.includes('条结果'));
                    const el2Text = el2 ? el2.textContent.trim() : 'none';
                    return {url: url.substring(0, 80), count_el: elText, fallback_el: el2Text};
                }
            """)
            print(f"  [i] 检索后URL: {debug_info.get('url', '?')}")
            print(f"  [i] 计数元素: {debug_info.get('count_el', '?')} | 备用: {debug_info.get('fallback_el', '?')}")
            count_text = debug_info.get('count_el') if debug_info.get('count_el') != 'none' else debug_info.get('fallback_el', '0')
            nums = re.findall(r'[\d,]+', count_text or '0')
            count = int(nums[0].replace(',', '')) if nums else 0
            return count
        except Exception:
            return 1  # 无法解析数量时假设有结果

    async def _sort_by_citations(self):
        """点击结果页的被引排序按钮（CNKI 用 <li> 元素，不导航离开 SPA）"""
        await asyncio.sleep(self.delay)
        try:
            # 等待排序按钮出现（<li> 或 <a>，文字为"被引"）
            try:
                await self.page.wait_for_function("""
                    () => [...document.querySelectorAll('li, a, span')].some(
                        el => el.textContent.trim() === '被引' || el.textContent.trim() === '被引量'
                    )
                """, timeout=20000)
            except Exception:
                print("  [!] 等待被引按钮超时，尝试直接点击...")

            result = await self.page.evaluate("""
                () => {
                    // CNKI 排序按钮是 <li> 元素（含 class 如 DESC/ASC order）
                    const items = [...document.querySelectorAll('li, a, span')];
                    const citeItem = items.find(el => {
                        const t = el.textContent.trim();
                        if (t !== '被引' && t !== '被引量') return false;
                        // 排除搜索表单内的元素
                        if (el.closest('.search-input, .input-row, .search-form, .btn-search-wrap')) return false;
                        return true;
                    });
                    if (citeItem) {
                        citeItem.click();
                        return 'clicked: ' + citeItem.tagName + ':' + citeItem.textContent.trim() +
                               ' parent=' + (citeItem.parentElement ? citeItem.parentElement.className.substring(0,30) : '');
                    }
                    // 诊断：列出所有短文字 li/a
                    const sample = items.filter(el => {
                        const t = el.textContent.trim();
                        return t.length > 0 && t.length < 8;
                    }).slice(0, 30).map(el => el.textContent.trim()).join(' | ');
                    return 'not found | sample: ' + sample.substring(0, 200);
                }
            """)
            print(f"  [i] 排序结果: {result[:200]}")

            await asyncio.sleep(self.delay * 2)
            try:
                await self.page.wait_for_load_state("networkidle", timeout=8000)
            except Exception:
                await asyncio.sleep(2)
        except Exception as e:
            print(f"  [!] 排序时出错: {e}，继续执行...")

    async def _set_50_per_page(self):
        """切换每页显示50条"""
        await asyncio.sleep(self.delay)
        try:
            # 等待 #perPageDiv 出现（排序后页面可能重新渲染）
            try:
                await self.page.wait_for_selector('#perPageDiv', state='visible', timeout=15000)
            except Exception:
                print("  [!] 等待 #perPageDiv 超时，尝试继续...")

            # #perPageDiv 是下拉组件：先点击 .sort-default 打开，再点击 li[data-val="50"]
            result = await self.page.evaluate("""
                () => {
                    const perDiv = document.querySelector('#perPageDiv');
                    if (!perDiv || perDiv.offsetParent === null)
                        return 'perPageDiv not found or hidden';

                    // 已经是50条则跳过
                    const curSpan = perDiv.querySelector('.sort-default span');
                    if (curSpan && curSpan.textContent.trim() === '50')
                        return 'already 50';

                    // 打开下拉
                    const sortDefault = perDiv.querySelector('.sort-default');
                    if (sortDefault) sortDefault.click();

                    // 直接 JS click 50 选项（绕过 display:none）
                    const li50 = perDiv.querySelector('li[data-val="50"] a');
                    if (li50) { li50.click(); return 'clicked50'; }

                    const allOptions = [...perDiv.querySelectorAll('li')].map(e => e.dataset.val).join('|');
                    return 'li[data-val=50] not found | options: ' + allOptions;
                }
            """)
            print(f"  [i] 每页50条: {result[:120]}")

            # 等待页面用50条重新渲染：networkidle + 固定延迟
            try:
                await self.page.wait_for_load_state("networkidle", timeout=12000)
            except Exception:
                pass
            await asyncio.sleep(3)
            print("  [i] 每页50条等待完成")
        except Exception as e:
            print(f"  [!] 切换每页50条时出错: {e}，继续执行...")

    async def _select_papers(self, max_count: int) -> int:
        """逐页全选，最多选 max_count 篇，返回实际选中数量"""
        total_selected = 0
        pages_needed = (max_count + 49) // 50  # 需要多少页

        for page_num in range(pages_needed):
            if page_num > 0:
                # 翻到下一页
                try:
                    next_btn = self.page.locator('.btn-next, a:has-text("下一页"), #PageNext').first
                    if not await next_btn.is_visible(timeout=3000):
                        break
                    await next_btn.click()
                    await asyncio.sleep(self.delay * 2)
                except Exception:
                    break

            # 全选当前页
            clicked = await self.page.evaluate("""
                () => {
                    const cb = document.querySelector('#selectCheckAll1');
                    if (cb) { cb.click(); return true; }
                    // 备用选择器
                    const cb2 = document.querySelector('input[name="selectCheckAll"]');
                    if (cb2) { cb2.click(); return true; }
                    return false;
                }
            """)
            await asyncio.sleep(self.delay)

            # 读取当前选中数量
            count_text = await self.page.evaluate("""
                () => {
                    const el = document.querySelector('#selectCount, .select-count');
                    return el ? el.textContent.trim() : '0';
                }
            """)
            try:
                selected = int(re.search(r'\d+', count_text).group())
            except Exception:
                selected = total_selected + 50

            total_selected = selected
            print(f"       第{page_num + 1}页全选完成，累计已选: {total_selected}")

            if total_selected >= max_count:
                break

        return total_selected

    async def _select_and_export_by_page(self, max_count: int) -> list[dict]:
        """逐页选中并导出（每页默认20条，与CNKI导出页面每次限制对齐）。"""
        results: list[dict] = []
        seen_titles: set[str] = set()
        total_collected = 0
        pages_needed = (min(max_count, 100) + 19) // 20  # 每页20条

        for page_num in range(pages_needed):
            if total_collected >= max_count:
                break

            if page_num > 0:
                # 翻到下一页
                try:
                    next_btn = self.page.locator('#Page_next_top, .btn-next, a:has-text("下一页")').first
                    if not await next_btn.is_visible(timeout=3000):
                        print(f"       没有第{page_num + 1}页，停止翻页")
                        break
                    await next_btn.click()
                    await asyncio.sleep(self.delay * 2)
                    try:
                        await self.page.wait_for_load_state("networkidle", timeout=8000)
                    except Exception:
                        await asyncio.sleep(2)
                except Exception:
                    break

            # 等待全选复选框就绪（超时后重试一次，给排序后页面更多渲染时间）
            checkbox_ready = False
            for cb_attempt in range(2):
                try:
                    await self.page.wait_for_selector('#selectCheckAll1', state='visible', timeout=15000)
                    checkbox_ready = True
                    break
                except Exception:
                    if cb_attempt == 0:
                        print(f"  [!] 等待 #selectCheckAll1 超时（第{page_num + 1}页），等待10秒后重试...")
                        await asyncio.sleep(10)
                    else:
                        print(f"  [!] #selectCheckAll1 二次超时（第{page_num + 1}页），停止")

            # 等待实际结果行渲染（避免排序刷新期间提前点击导致选中数为0）
            try:
                await self.page.wait_for_function("""
                    () => {
                        const items = document.querySelectorAll('input[name="CookieName"], tr[data-uid], .result-item');
                        return items.length > 0;
                    }
                """, timeout=10000)
                await asyncio.sleep(0.5)
            except Exception:
                await asyncio.sleep(2)

            # 清除上一批选择（避免跨页累计）
            await self.page.evaluate("""
                () => {
                    const a = document.querySelector('a[href*="filenameClear"]');
                    if (a) a.click();
                }
            """)
            await asyncio.sleep(0.5)

            # 全选当前页
            await self.page.evaluate("""
                () => {
                    const cb = document.querySelector('#selectCheckAll1');
                    if (cb) { cb.click(); return; }
                    const cb2 = document.querySelector('input[name="selectCheckAll"]');
                    if (cb2) cb2.click();
                }
            """)
            await asyncio.sleep(self.delay + 0.5)

            # 读取选中数量
            count_text = await self.page.evaluate("""
                () => {
                    const el = document.querySelector('#selectCount, .select-count');
                    return el ? el.textContent.trim() : '0';
                }
            """)
            try:
                selected = int(re.search(r'\d+', count_text).group())
            except Exception:
                selected = 0

            if selected == 0:
                print(f"       第{page_num + 1}页无选中项，停止")
                break

            print(f"       第{page_num + 1}页已选 {selected} 篇，开始导出...")
            try:
                articles = await self._export_and_parse()
            except CNKIParseError as e:
                print(f"  [!] 第{page_num + 1}页解析失败: {e}，跳过")
                articles = []
            except Exception as e:
                print(f"  [!] 第{page_num + 1}页导出失败: {e}，停止")
                break

            new_count = 0
            for a in articles:
                title = a.get('title', '').strip()
                if title and title not in seen_titles:
                    seen_titles.add(title)
                    results.append(a)
                    new_count += 1
                    total_collected += 1
            print(f"       第{page_num + 1}页导出 {len(articles)} 篇，新增 {new_count} 篇（累计 {total_collected} 篇）")

            # 页间冷却（防止频率限制）
            if self.batch_sleep > 0 and page_num < pages_needed - 1:
                print(f"  [~] 冷却 {self.batch_sleep} 秒，防止触发频率限制...")
                await asyncio.sleep(self.batch_sleep)

            # 清除选择，准备下一页
            await self.page.evaluate("""
                () => {
                    const a = document.querySelector('a[href*="filenameClear"]');
                    if (a) a.click();
                }
            """)
            await asyncio.sleep(1)

        return results

    async def _close_stale_export_tabs(self):
        """关闭浏览器中所有残留的 export.html 标签，避免累积触发频率限制"""
        try:
            for p in self.page.context.pages:
                if p != self.page and 'export.html' in p.url:
                    try:
                        await p.close()
                    except Exception:
                        pass
        except Exception:
            pass

    async def _export_and_parse(self) -> list[dict]:
        """触发导出、切换到导出页、提取数据、解析返回"""
        # 每次导出前先清理残留的 export 标签，防止累积触发频率限制
        await self._close_stale_export_tabs()
        # 检查登录
        await self._ensure_logged_in()

        # 使用 expect_popup 捕获 $.PostWindow() 打开的新窗口
        export_page = None
        for attempt in range(2):
            try:
                async with self.page.expect_popup(timeout=15000) as popup_info:
                    result = await self.page.evaluate("""
                        () => {
                            if (typeof $ === 'undefined' || typeof $.filenameGet !== 'function') {
                                return {error: 'jQuery or CNKI functions not available'};
                            }
                            const filename = $.filenameGet();
                            const searchinfo = $.searchinfoGet();
                            const mapIndex = $.indexGet();
                            const baseUrl = document.querySelector('#hidDocumentManageUrl')?.value
                                            || 'https://kns.cnki.net/dm8';
                            const exportUrl = baseUrl + '/manage/export.html?language=CHS&uniplatform=NZKPT';
                            $.PostWindow(exportUrl, {
                                displaymode: 'NEW',
                                filename: filename,
                                searchinfo: searchinfo,
                                mapIndex: mapIndex
                            });
                            return {exportUrl, filename: String(filename).substring(0, 50)};
                        }
                    """)
                export_page = await popup_info.value
                print(f"       导出页面已打开: {export_page.url[:80]}")
                break
            except Exception as e:
                if attempt == 0:
                    print(f"  [!] 导出第一次尝试失败: {e}，等待后重试...")
                    await asyncio.sleep(3)
                else:
                    raise CNKIExportFailed(f"无法打开导出页面: {e}")

        # 检查是否跳到了登录页
        if export_page is None:
            raise CNKIExportFailed("导出页面未打开")

        await export_page.wait_for_load_state("domcontentloaded", timeout=15000)
        await asyncio.sleep(1.5)

        # chrome-error 表示 $.filenameGet() 返回空或 POST 数据无效，关闭并重试一次
        if "chrome-error://" in export_page.url:
            print("  [!] 导出 popup 为 chrome-error，等待 3 秒后重试...")
            try:
                await export_page.close()
            except Exception:
                pass
            await asyncio.sleep(3)
            # 重试：重新触发 $.PostWindow()
            try:
                async with self.page.expect_popup(timeout=15000) as popup_info2:
                    await self.page.evaluate("""
                        () => {
                            const filename = $.filenameGet();
                            const searchinfo = $.searchinfoGet();
                            const mapIndex = $.indexGet();
                            const baseUrl = document.querySelector('#hidDocumentManageUrl')?.value
                                            || 'https://kns.cnki.net/dm8';
                            const exportUrl = baseUrl + '/manage/export.html?language=CHS&uniplatform=NZKPT';
                            $.PostWindow(exportUrl, {
                                displaymode: 'NEW',
                                filename: filename,
                                searchinfo: searchinfo,
                                mapIndex: mapIndex
                            });
                        }
                    """)
                export_page = await popup_info2.value
                await export_page.wait_for_load_state("domcontentloaded", timeout=15000)
                await asyncio.sleep(1.5)
                print(f"       重试后导出页面: {export_page.url[:80]}")
            except Exception as e2:
                raise CNKIExportFailed(f"重试导出也失败: {e2}")

        if "member.cnki.net" in export_page.url or "login" in export_page.url.lower():
            await export_page.close()
            print("\n[!] 导出时跳转到登录页，请在浏览器中登录知网（自动等待 30 秒后重试）...")
            await asyncio.sleep(30)
            raise CNKILoginRequired("导出时未登录")

        # 在导出页选择"查新（引文格式）"
        await self._select_citation_format(export_page)

        # 等待内容加载，检测验证码（用DOM可见性判断，避免页脚静态文本误触发）
        await asyncio.sleep(2)
        captcha_printed = False
        for attempt in range(90):  # 最多等 180 秒
            # 用 offsetParent 判断验证码组件是否真实可见（页脚的"安全验证"文字不会触发）
            captcha_active = await export_page.evaluate("""
                () => {
                    const sels = [
                        '.tencent-captcha-dy__content',
                        '#tCaptchaDyContent',
                        '.blockPuzzle',
                        '.geetest_widget',
                        '.nc-container',
                        '#slideBg',
                        '.tcaptcha-transform'
                    ];
                    for (const sel of sels) {
                        const el = document.querySelector(sel);
                        if (el && el.offsetParent !== null) return true;
                    }
                    return false;
                }
            """)
            if captcha_active:
                if not captcha_printed:
                    print("\n[!] 导出页面出现滑块验证，请在 Chrome 窗口中完成拼图验证（自动等待，最多 180 秒）...")
                    captcha_printed = True
                await asyncio.sleep(2)
                continue
            # 验证码已消失或不存在，继续导出
            if captcha_printed:
                print("  [✓] 验证码已通过，继续导出...")
                await self._select_citation_format(export_page)
                await asyncio.sleep(2)
            break
        else:
            raise CNKIExportFailed("导出页面验证码等待超时（180秒），请手动刷新或稍后重试")

        # 提取文本并解析，无论成功与否都关闭 popup
        try:
            raw_text = await export_page.inner_text('body')
            articles = parse_citation_text(raw_text)
            if not articles:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                raw_path = OUTPUT_DIR / f"cnki_raw_export_{ts}.txt"
                raw_path.write_text(raw_text, encoding="utf-8")
                raise CNKIParseError(f"解析失败，原始数据已保存到: {raw_path}")
            return articles
        finally:
            # 无论成功、解析失败还是异常，都关闭 popup 防止标签堆积
            try:
                await export_page.close()
            except Exception:
                pass
            await asyncio.sleep(1)

    async def _select_citation_format(self, export_page: Page):
        """在导出页面点击'查新（引文格式）'"""
        try:
            format_btn = export_page.locator('a:has-text("查新"), li:has-text("查新（引文格式）"), .format-item:has-text("查新")').first
            if await format_btn.is_visible(timeout=5000):
                await format_btn.click()
                await asyncio.sleep(2)
                return
        except Exception:
            pass

        # 备用：evaluate
        await export_page.evaluate("""
            () => {
                const items = [...document.querySelectorAll('a, li, span, div')];
                const item = items.find(i => i.textContent.includes('查新') && i.textContent.includes('引文'));
                if (item) { item.click(); return true; }
                // 也尝试只包含"查新"的选项
                const item2 = items.find(i => i.textContent.trim() === '查新（引文格式）');
                if (item2) { item2.click(); return true; }
                return false;
            }
        """)
        await asyncio.sleep(2)


# ── 数据解析 ──────────────────────────────────────────────

def parse_citation_text(raw_text: str) -> list[dict]:
    """
    将知网查新引文格式文本解析为字典列表。
    格式：[1]\n作者. 标题[J]. 期刊, 年份, 卷(期): 页码.\n摘要: ...\n
    """
    start = raw_text.find('[1]\n')
    if start < 0:
        return []

    body = raw_text[start:]
    # 按 [N]\n 分割，保留分隔符中的数字
    parts = re.split(r'\[(\d+)\]\n', body)

    articles = []
    for i in range(1, len(parts), 2):
        try:
            num = int(parts[i])
            content = parts[i + 1].strip() if i + 1 < len(parts) else ''
            if not content:
                continue

            # 分离引文行与摘要
            abs_match = re.search(r'\n摘要[：:]', content)
            if abs_match:
                citation_line = content[:abs_match.start()].strip()
                abstract = content[abs_match.start():].strip()
                abstract = re.sub(r'^摘要[：:]', '', abstract).strip()
            else:
                citation_line = content.strip()
                abstract = ''

            # 解析引文行：作者. 标题[J]. 期刊信息
            authors, title, source, date = '', '', '', ''
            m = re.match(r'^(.*?)\.\s*(.*?)\[[A-Z/]+\]\.\s*(.*)$', citation_line, re.DOTALL)
            if m:
                authors = m.group(1).strip()
                title = m.group(2).strip()
                journal_info = m.group(3).strip().rstrip('.')
                j_match = re.match(r'^(.*?),\s*(\d{4})(.*?)$', journal_info)
                if j_match:
                    source = j_match.group(1).strip()
                    date = j_match.group(2) + j_match.group(3).strip()
                else:
                    source = journal_info
            else:
                title = citation_line  # 无法解析时用原文

            articles.append({
                'num': num,
                'title': title,
                'authors': authors,
                'source': source,
                'date': date,
                'abstract': abstract,
            })
        except Exception as e:
            print(f"  [!] 解析第 {i} 条时出错: {e}")

    return articles


# ── Excel 输出 ────────────────────────────────────────────

def save_to_excel(articles: list[dict], output_path: Path, keyword_summary: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "检索结果"

    headers = ["序号", "标题", "作者", "来源期刊", "发表时间", "摘要"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for i, article in enumerate(articles, 1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=article.get('title', ''))
        ws.cell(row=i + 1, column=3, value=article.get('authors', ''))
        ws.cell(row=i + 1, column=4, value=article.get('source', ''))
        ws.cell(row=i + 1, column=5, value=article.get('date', ''))
        cell = ws.cell(row=i + 1, column=6, value=article.get('abstract', ''))
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 80

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

    wb.save(output_path)


# ── 主函数 ────────────────────────────────────────────────

async def main():
    parser = argparse.ArgumentParser(description='知网CNKI高级检索自动化工具')
    parser.add_argument('--keywords', action='append', required=True,
                        metavar='KEYWORD_GROUP',
                        help='关键词组（可重复多次，每次为一组，同义词用" + "连接）')
    parser.add_argument('--max-results', type=int, default=100,
                        help='最多导出论文数量（默认100，最大100）')
    parser.add_argument('--port', type=int, default=9222,
                        help='Chrome CDP 调试端口（默认9222）')
    parser.add_argument('--min-groups', type=int, default=2,
                        help='自动降级的最少保留组数（默认2，不会自动降到1组）')
    parser.add_argument('--delay', type=int, default=1000,
                        help='操作间延迟（毫秒，默认1000）')
    parser.add_argument('--batch-sleep', type=int, default=0,
                        help='每页导出后的冷却时间（秒，默认0），连续多批检索时建议设为30-120')
    parser.add_argument('--output-dir', type=str, default=str(OUTPUT_DIR),
                        help='输出目录（默认~/Downloads/）')
    args = parser.parse_args()

    max_results = min(args.max_results, 100)
    output_dir = Path(args.output_dir).expanduser()
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n===== 知网高级检索自动化工具 =====")
    print(f"关键词分组: {args.keywords}")
    print(f"最大结果数: {max_results}")
    print(f"输出目录: {output_dir}")
    print(f"===================================\n")

    async with async_playwright() as p:
        browser = None
        connected_via_cdp = False

        # 优先连接现有 Chrome（保留登录会话）
        try:
            browser = await p.chromium.connect_over_cdp(f"http://localhost:{args.port}")
            existing_context = browser.contexts[0] if browser.contexts else None
            if existing_context is None:
                raise Exception("No browser context found")
            context = existing_context
            connected_via_cdp = True
            print(f"[✓] 已连接到现有 Chrome（端口 {args.port}）")
        except Exception as e:
            print(f"[!] 无法连接到 Chrome CDP（{e}），启动新浏览器...")
            browser = await p.chromium.launch(
                headless=False,
                args=["--disable-blink-features=AutomationControlled"]
            )
            context = await browser.new_context(
                locale="zh-CN",
                viewport={"width": 1280, "height": 900},
                ignore_https_errors=True,
                user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                           "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            )
            print("[✓] 新浏览器已启动（有头模式）")
            print("[!] 提示：如需使用已登录的知网账号，请通过 Chrome 调试模式连接。")
            print("    启动命令：open -a 'Google Chrome' --args --remote-debugging-port=9222\n")

        # 始终新开标签页，不干扰已有页面
        page = await context.new_page()

        try:
            searcher = CNKISearcher(page, delay_ms=args.delay, batch_sleep_s=args.batch_sleep)
            articles = await searcher.run(args.keywords, max_results=max_results,
                                          min_groups=args.min_groups)

            # 生成文件名
            keyword_summary = "_".join(args.keywords[0].split()[:2])[:20]
            # 清理文件名中的特殊字符
            keyword_summary = re.sub(r'[^\w\u4e00-\u9fff]', '_', keyword_summary)
            date_str = datetime.now().strftime("%Y%m%d_%H%M")
            filename = f"知网检索_{keyword_summary}_{date_str}.xlsx"
            output_path = output_dir / filename

            save_to_excel(articles, output_path, keyword_summary)

            print(f"\n[✓] 检索完成！共提取 {len(articles)} 篇论文")
            print(f"[✓] 文件已保存：{output_path}")

        except CNKINoResults as e:
            print(f"\n[!] {e}")
            sys.exit(1)
        except CNKIExportFailed as e:
            print(f"\n[✗] 导出失败: {e}")
            sys.exit(2)
        except CNKIParseError as e:
            print(f"\n[✗] 数据解析失败: {e}")
            sys.exit(3)
        except KeyboardInterrupt:
            print("\n[!] 用户中断")
            sys.exit(0)
        finally:
            if connected_via_cdp:
                await page.close()  # 关闭新开的标签页，不影响其他已有页面
            else:
                await browser.close()


if __name__ == "__main__":
    asyncio.run(main())
