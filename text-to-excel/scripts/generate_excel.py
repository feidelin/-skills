#!/usr/bin/env python3
"""
Text-to-Excel Generator
Generates .xlsx files from structured JSON data with full formatting support.

Usage:
    python3 generate_excel.py <json_config_path> <output_path>

JSON config format: see references/config_schema.md
"""

import json
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    BarChart, LineChart, PieChart, AreaChart, ScatterChart, Reference
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import (
    CellIsRule, ColorScaleRule, DataBarRule
)


def apply_font(cell, font_cfg):
    """Apply font settings to a cell."""
    if not font_cfg:
        return
    cell.font = Font(
        name=font_cfg.get("name", "Calibri"),
        size=font_cfg.get("size", 11),
        bold=font_cfg.get("bold", False),
        italic=font_cfg.get("italic", False),
        color=font_cfg.get("color"),
        underline="single" if font_cfg.get("underline") else None,
        strike=font_cfg.get("strikethrough", False),
    )


def apply_fill(cell, fill_cfg):
    """Apply fill/background settings to a cell."""
    if not fill_cfg:
        return
    cell.fill = PatternFill(
        start_color=fill_cfg.get("color", "FFFFFF"),
        end_color=fill_cfg.get("color", "FFFFFF"),
        fill_type=fill_cfg.get("type", "solid"),
    )


def apply_alignment(cell, align_cfg):
    """Apply alignment settings to a cell."""
    if not align_cfg:
        return
    cell.alignment = Alignment(
        horizontal=align_cfg.get("horizontal", "general"),
        vertical=align_cfg.get("vertical", "center"),
        wrap_text=align_cfg.get("wrap_text", False),
        text_rotation=align_cfg.get("text_rotation", 0),
        indent=align_cfg.get("indent", 0),
    )


def apply_border(cell, border_cfg):
    """Apply border settings to a cell."""
    if not border_cfg:
        return
    sides = {}
    for side_name in ["left", "right", "top", "bottom"]:
        side_cfg = border_cfg.get(side_name)
        if side_cfg:
            sides[side_name] = Side(
                style=side_cfg.get("style", "thin"),
                color=side_cfg.get("color", "000000"),
            )
        elif border_cfg.get("all"):
            all_cfg = border_cfg["all"]
            sides[side_name] = Side(
                style=all_cfg.get("style", "thin"),
                color=all_cfg.get("color", "000000"),
            )
    cell.border = Border(**sides)


def apply_number_format(cell, fmt):
    """Apply number format to a cell."""
    if not fmt:
        return
    fmt_map = {
        "number": "#,##0",
        "decimal2": "#,##0.00",
        "percent": "0.00%",
        "currency_usd": "$#,##0.00",
        "currency_cny": "¥#,##0.00",
        "currency_eur": "€#,##0.00",
        "date": "YYYY-MM-DD",
        "datetime": "YYYY-MM-DD HH:MM:SS",
        "time": "HH:MM:SS",
        "text": "@",
    }
    if fmt in fmt_map:
        cell.number_format = fmt_map[fmt]
    else:
        # Allow custom format strings
        cell.number_format = fmt


def apply_cell_style(cell, style_cfg):
    """Apply all style settings to a cell."""
    if not style_cfg:
        return
    apply_font(cell, style_cfg.get("font"))
    apply_fill(cell, style_cfg.get("fill"))
    apply_alignment(cell, style_cfg.get("alignment"))
    apply_border(cell, style_cfg.get("border"))
    apply_number_format(cell, style_cfg.get("number_format"))


def apply_default_header_style(cell):
    """Apply default header style if no custom style is specified."""
    cell.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )


def apply_default_data_style(cell, row_idx):
    """Apply default alternating row style."""
    cell.font = Font(size=11, name="Calibri")
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    if row_idx % 2 == 0:
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def cast_cell_value(value, col_type=None):
    """Cast cell value to appropriate Python type."""
    if value is None or value == "":
        return value
    if col_type == "number":
        try:
            if isinstance(value, str) and "." in value:
                return float(value)
            return int(value) if isinstance(value, str) else value
        except (ValueError, TypeError):
            return value
    if col_type == "float":
        try:
            return float(value)
        except (ValueError, TypeError):
            return value
    if col_type == "bool":
        if isinstance(value, str):
            return value.lower() in ("true", "yes", "1", "是")
        return bool(value)
    return value


def add_chart(ws, chart_cfg, data_start_row, data_end_row):
    """Add a chart to the worksheet."""
    chart_type = chart_cfg.get("type", "bar")
    title = chart_cfg.get("title", "")
    x_col = chart_cfg.get("x_column", 1)
    y_cols = chart_cfg.get("y_columns", [2])
    position = chart_cfg.get("position", "E2")
    width = chart_cfg.get("width", 15)
    height = chart_cfg.get("height", 10)

    chart_classes = {
        "bar": BarChart,
        "line": LineChart,
        "pie": PieChart,
        "area": AreaChart,
        "scatter": ScatterChart,
    }
    ChartClass = chart_classes.get(chart_type, BarChart)
    chart = ChartClass()
    chart.title = title
    chart.width = width
    chart.height = height

    if chart_cfg.get("x_title"):
        chart.x_axis.title = chart_cfg["x_title"]
    if chart_cfg.get("y_title"):
        chart.y_axis.title = chart_cfg["y_title"]

    cats = Reference(ws, min_col=x_col, min_row=data_start_row, max_row=data_end_row)

    for y_col in y_cols:
        data = Reference(ws, min_col=y_col, min_row=data_start_row - 1, max_row=data_end_row)
        if chart_type == "scatter":
            x_data = Reference(ws, min_col=x_col, min_row=data_start_row, max_row=data_end_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(x_data)
        else:
            chart.add_data(data, titles_from_data=True)

    if chart_type != "scatter":
        chart.set_categories(cats)

    ws.add_chart(chart, position)


def add_data_validation(ws, dv_cfg):
    """Add data validation to a range."""
    dv_type = dv_cfg.get("type", "list")
    formula = dv_cfg.get("formula", "")
    cell_range = dv_cfg.get("range", "A1")
    prompt_title = dv_cfg.get("prompt_title", "")
    prompt_msg = dv_cfg.get("prompt_message", "")
    error_title = dv_cfg.get("error_title", "Invalid")
    error_msg = dv_cfg.get("error_message", "Please enter a valid value.")

    if dv_type == "list":
        dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
    elif dv_type == "whole":
        min_val = dv_cfg.get("min", 0)
        max_val = dv_cfg.get("max", 100)
        dv = DataValidation(type="whole", operator="between",
                            formula1=str(min_val), formula2=str(max_val))
    elif dv_type == "decimal":
        min_val = dv_cfg.get("min", 0.0)
        max_val = dv_cfg.get("max", 100.0)
        dv = DataValidation(type="decimal", operator="between",
                            formula1=str(min_val), formula2=str(max_val))
    elif dv_type == "date":
        dv = DataValidation(type="date", operator="between",
                            formula1=dv_cfg.get("min", "2000-01-01"),
                            formula2=dv_cfg.get("max", "2099-12-31"))
    elif dv_type == "textLength":
        dv = DataValidation(type="textLength", operator="lessThanOrEqual",
                            formula1=str(dv_cfg.get("max", 255)))
    else:
        dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)

    if prompt_title:
        dv.prompt = prompt_msg
        dv.promptTitle = prompt_title
        dv.showInputMessage = True
    if error_title:
        dv.error = error_msg
        dv.errorTitle = error_title
        dv.showErrorMessage = True

    dv.add(cell_range)
    ws.add_data_validation(dv)


def add_conditional_formatting(ws, cf_cfg):
    """Add conditional formatting to a range."""
    cf_type = cf_cfg.get("type", "cell_is")
    cell_range = cf_cfg.get("range", "A1:A100")

    if cf_type == "cell_is":
        operator = cf_cfg.get("operator", "greaterThan")
        formula = cf_cfg.get("formula", "0")
        font_color = cf_cfg.get("font_color")
        fill_color = cf_cfg.get("fill_color", "FFC7CE")
        font = Font(color=font_color) if font_color else None
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        rule = CellIsRule(operator=operator, formula=[str(formula)], font=font, fill=fill)
        ws.conditional_formatting.add(cell_range, rule)
    elif cf_type == "color_scale":
        start_color = cf_cfg.get("start_color", "F8696B")
        mid_color = cf_cfg.get("mid_color", "FFEB84")
        end_color = cf_cfg.get("end_color", "63BE7B")
        rule = ColorScaleRule(
            start_type="min", start_color=start_color,
            mid_type="percentile", mid_value=50, mid_color=mid_color,
            end_type="max", end_color=end_color,
        )
        ws.conditional_formatting.add(cell_range, rule)
    elif cf_type == "data_bar":
        color = cf_cfg.get("color", "638EC6")
        rule = DataBarRule(start_type="min", end_type="max", color=color)
        ws.conditional_formatting.add(cell_range, rule)


def process_sheet(wb, sheet_cfg, is_first=False):
    """Process a single sheet configuration."""
    sheet_name = sheet_cfg.get("name", "Sheet1")
    if is_first:
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)

    headers = sheet_cfg.get("headers", [])
    data = sheet_cfg.get("data", [])
    column_types = sheet_cfg.get("column_types", [])
    column_widths = sheet_cfg.get("column_widths", [])
    header_style = sheet_cfg.get("header_style")
    data_style = sheet_cfg.get("data_style")
    merge_cells = sheet_cfg.get("merge_cells", [])
    freeze_pane = sheet_cfg.get("freeze_pane")
    auto_filter = sheet_cfg.get("auto_filter", False)
    row_heights = sheet_cfg.get("row_heights", {})
    charts = sheet_cfg.get("charts", [])
    data_validations = sheet_cfg.get("data_validations", [])
    conditional_formats = sheet_cfg.get("conditional_formats", [])
    title_row = sheet_cfg.get("title_row")
    auto_width = sheet_cfg.get("auto_width", True)

    current_row = 1

    # Title row (optional, spanning all columns)
    if title_row:
        title_text = title_row.get("text", "")
        title_style = title_row.get("style", {})
        ncols = len(headers) if headers else 1
        cell = ws.cell(row=current_row, column=1, value=title_text)
        if ncols > 1:
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=ncols)
        default_title_style = {
            "font": {"bold": True, "size": 14, "name": "Calibri"},
            "alignment": {"horizontal": "center", "vertical": "center"},
        }
        merged_style = {**default_title_style, **title_style}
        apply_cell_style(cell, merged_style)
        if title_row.get("height"):
            ws.row_dimensions[current_row].height = title_row["height"]
        else:
            ws.row_dimensions[current_row].height = 30
        current_row += 1

    # Header row
    header_start_row = current_row
    if headers:
        for col_idx, header in enumerate(headers, 1):
            if isinstance(header, dict):
                cell = ws.cell(row=current_row, column=col_idx, value=header.get("text", ""))
                apply_cell_style(cell, header.get("style"))
            else:
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                if header_style:
                    apply_cell_style(cell, header_style)
                else:
                    apply_default_header_style(cell)
        current_row += 1

    # Data rows
    data_start_row = current_row
    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data, 1):
            col_type = column_types[col_idx - 1] if col_idx - 1 < len(column_types) else None
            casted = cast_cell_value(value, col_type)
            cell = ws.cell(row=current_row, column=col_idx, value=casted)
            if data_style:
                apply_cell_style(cell, data_style)
            else:
                apply_default_data_style(cell, row_idx)
            # Apply column-specific number format
            if col_type:
                fmt_map = {
                    "percent": "0.00%",
                    "currency_usd": "$#,##0.00",
                    "currency_cny": "¥#,##0.00",
                    "date": "YYYY-MM-DD",
                }
                if col_type in fmt_map:
                    cell.number_format = fmt_map[col_type]
        current_row += 1
    data_end_row = current_row - 1

    # Column widths
    if column_widths:
        for col_idx, width in enumerate(column_widths, 1):
            if width:
                ws.column_dimensions[get_column_letter(col_idx)].width = width
    elif auto_width and headers:
        for col_idx, header in enumerate(headers, 1):
            header_text = header.get("text", header) if isinstance(header, dict) else header
            max_len = len(str(header_text))
            for row_data in data:
                if col_idx - 1 < len(row_data):
                    val_len = len(str(row_data[col_idx - 1])) if row_data[col_idx - 1] is not None else 0
                    max_len = max(max_len, val_len)
            # Approx: each CJK char ~2 width, ASCII ~1.2 width
            estimated = max_len * 1.5 + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(estimated, 10), 60)

    # Row heights
    for row_str, height in row_heights.items():
        ws.row_dimensions[int(row_str)].height = height

    # Merge cells
    for merge in merge_cells:
        if isinstance(merge, str):
            ws.merge_cells(merge)
        elif isinstance(merge, dict):
            ws.merge_cells(
                start_row=merge["start_row"],
                start_column=merge["start_col"],
                end_row=merge["end_row"],
                end_column=merge["end_col"],
            )
            # Apply style to merged area if provided
            if merge.get("value") is not None:
                cell = ws.cell(row=merge["start_row"], column=merge["start_col"],
                               value=merge["value"])
                if merge.get("style"):
                    apply_cell_style(cell, merge["style"])

    # Freeze pane
    if freeze_pane:
        ws.freeze_panes = freeze_pane

    # Auto filter
    if auto_filter and headers:
        ws.auto_filter.ref = f"A{header_start_row}:{get_column_letter(len(headers))}{data_end_row}"

    # Charts
    for chart_cfg in charts:
        add_chart(ws, chart_cfg, data_start_row, data_end_row)

    # Data validations
    for dv_cfg in data_validations:
        add_data_validation(ws, dv_cfg)

    # Conditional formatting
    for cf_cfg in conditional_formats:
        add_conditional_formatting(ws, cf_cfg)

    return ws


def generate_excel(config, output_path):
    """Generate Excel file from config dict."""
    wb = Workbook()
    sheets = config.get("sheets", [])
    if not sheets:
        # Single-sheet mode: treat entire config as one sheet
        sheets = [config]

    for idx, sheet_cfg in enumerate(sheets):
        process_sheet(wb, sheet_cfg, is_first=(idx == 0))

    # Global settings
    if config.get("print_settings"):
        ps = config["print_settings"]
        for ws in wb.worksheets:
            if ps.get("orientation") == "landscape":
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            if ps.get("paper_size"):
                ws.page_setup.paperSize = ps["paper_size"]
            if ps.get("fit_to_page"):
                ws.page_setup.fitToPage = True
                ws.page_setup.fitToWidth = ps.get("fit_to_width", 1)
                ws.page_setup.fitToHeight = ps.get("fit_to_height", 0)

    wb.save(output_path)
    print(f"Excel file saved: {output_path}")
    return output_path


def main():
    if len(sys.argv) < 3:
        print("Usage: python3 generate_excel.py <json_config_path> <output_path>")
        sys.exit(1)

    config_path = sys.argv[1]
    output_path = sys.argv[2]

    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)

    generate_excel(config, output_path)


if __name__ == "__main__":
    main()
