#!/usr/bin/env python3
"""
OpenAlex 外文学术文献检索脚本
- 支持多关键词组 AND 联合检索（各组内 OR，组间取交集）
- 自动过滤社会科学领域
- 按被引量排序，解码摘要，输出 Excel
- 支持期刊分区信息（JCR Q1-Q4 / 中科院分区 + 2年影响因子）
"""

import argparse
import json
import sys
import time
import urllib.request
import urllib.parse
from datetime import date
from pathlib import Path
from collections import Counter

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("[!] 请先安装: pip install openpyxl")
    sys.exit(1)

# 社会科学领域 concept IDs（用于过滤非社科文献）
SOCIAL_SCIENCE_CONCEPTS = [
    "C144024400",  # Sociology
    "C17744445",   # Political science
    "C162324750",  # Economics
    "C15744967",   # Psychology
    "C205649164",  # Geography
    "C142362112",  # Art
    "C95457728",   # History
    "C41008148",   # Computer science (数字社会学需要)
]

# 精选社科期刊分区表（journal name小写 → (JCR分区, 中科院分区)）
# JCR分区基于社会科学引文索引各学科分类；中科院分区按JIF百分位
JOURNAL_RANKING_TABLE = {
    # ── 社会学顶刊 ──────────────────────────────────────────────────────────
    "american sociological review":                       ("Q1", "1区"),
    "american journal of sociology":                      ("Q1", "1区"),
    "annual review of sociology":                         ("Q1", "1区"),
    "social forces":                                      ("Q1", "1区"),
    "sociological theory":                                ("Q1", "2区"),
    "sociological methods & research":                    ("Q1", "2区"),
    "sociological methods and research":                  ("Q1", "2区"),
    "theory and society":                                 ("Q1", "2区"),
    "british journal of sociology":                       ("Q1", "2区"),
    "sociology":                                          ("Q1", "2区"),
    "current sociology":                                  ("Q2", "2区"),
    "social problems":                                    ("Q2", "2区"),
    "qualitative sociology":                              ("Q2", "3区"),
    "gender & society":                                   ("Q1", "2区"),
    "gender and society":                                 ("Q1", "2区"),
    "social science research":                            ("Q2", "2区"),
    "european sociological review":                       ("Q1", "2区"),
    "acta sociologica":                                   ("Q2", "3区"),
    "sociological review":                                ("Q2", "3区"),
    "the sociological review":                            ("Q2", "3区"),
    "cultural sociology":                                 ("Q2", "3区"),
    "sociology of education":                             ("Q1", "2区"),
    "qualitative inquiry":                                ("Q2", "3区"),

    # ── 劳工/组织/管理 ───────────────────────────────────────────────────────
    "administrative science quarterly":                   ("Q1", "1区"),
    "organization science":                               ("Q1", "1区"),
    "organization studies":                               ("Q1", "1区"),
    "organization":                                       ("Q1", "2区"),
    "journal of management studies":                      ("Q1", "1区"),
    "journal of management":                              ("Q1", "1区"),
    "human relations":                                    ("Q1", "2区"),
    "work employment and society":                        ("Q1", "2区"),
    "work, employment and society":                       ("Q1", "2区"),
    "industrial and labor relations review":              ("Q1", "2区"),
    "british journal of industrial relations":            ("Q1", "2区"),
    "work and occupations":                               ("Q2", "3区"),
    "labor studies journal":                              ("Q3", "4区"),
    "journal of labor economics":                         ("Q1", "1区"),
    "journal of labor research":                          ("Q3", "4区"),
    "economic and industrial democracy":                  ("Q2", "3区"),
    "transfer: european review of labour and research":   ("Q3", "4区"),

    # ── 平台经济/数字劳动/零工经济 ──────────────────────────────────────────
    "new media & society":                                ("Q1", "1区"),
    "new media and society":                              ("Q1", "1区"),
    "information communication & society":                ("Q1", "2区"),
    "information, communication & society":               ("Q1", "2区"),
    "journal of computer-mediated communication":         ("Q1", "1区"),
    "social media + society":                             ("Q1", "2区"),
    "media, culture & society":                           ("Q2", "3区"),
    "media culture & society":                            ("Q2", "3区"),
    "big data & society":                                 ("Q1", "2区"),
    "information technology & people":                    ("Q2", "3区"),
    "journal of information technology":                  ("Q1", "1区"),
    "mis quarterly":                                      ("Q1", "1区"),
    "global networks":                                    ("Q1", "2区"),
    "economy and society":                                ("Q1", "2区"),
    "socio-economic review":                              ("Q1", "2区"),
    "journal of consumer research":                       ("Q1", "1区"),

    # ── 城市/移动性/空间 ─────────────────────────────────────────────────────
    "urban studies":                                      ("Q1", "1区"),
    "environment and planning a":                         ("Q1", "1区"),
    "environment and planning a: economy and space":      ("Q1", "1区"),
    "city":                                               ("Q1", "2区"),
    "international journal of urban and regional research":("Q1", "2区"),
    "journal of urban affairs":                           ("Q2", "3区"),
    "mobilities":                                         ("Q2", "2区"),
    "transportation research part a: policy and practice":("Q1", "1区"),
    "transport policy":                                   ("Q1", "2区"),
    "travel behaviour and society":                       ("Q2", "3区"),
}

# ── 根据 2yr影响因子 估算分区（用于未收录期刊的兜底推断）────────────────────
def infer_quartile_from_impact(impact_factor):
    """2yr_mean_citedness 估算分区（社科领域经验阈值）"""
    if impact_factor is None:
        return ("", "")
    if impact_factor >= 3.0:
        return ("Q1*", "1-2区*")
    elif impact_factor >= 1.5:
        return ("Q2*", "2-3区*")
    elif impact_factor >= 0.8:
        return ("Q3*", "3-4区*")
    else:
        return ("Q4*", "4区*")
# * 号表示"根据影响因子估算，非官方分区"


def decode_abstract(inverted_index):
    """将 OpenAlex 的倒排索引格式摘要还原为普通文本"""
    if not inverted_index:
        return ""
    words = {}
    for word, positions in inverted_index.items():
        for pos in positions:
            words[pos] = word
    return " ".join(words[k] for k in sorted(words))


def openalex_get(endpoint, params, retries=3):
    """带重试的 OpenAlex API 请求（无需 API key）"""
    query_string = urllib.parse.urlencode(params)
    url = f"https://api.openalex.org/{endpoint}?{query_string}"

    for attempt in range(retries):
        try:
            req = urllib.request.Request(url)
            req.add_header("User-Agent", "AcademicResearchTool/1.0 (mailto:research@example.com)")
            with urllib.request.urlopen(req, timeout=30) as resp:
                return json.loads(resp.read().decode())
        except Exception as e:
            if attempt < retries - 1:
                print(f"  [!] 请求失败（{e}），{1.5*(attempt+1):.1f}s 后重试...")
                time.sleep(1.5 * (attempt + 1))
            else:
                print(f"  [✗] 请求最终失败: {url[:80]}")
                return None


def fetch_source_stats(source_ids):
    """
    批量查询 OpenAlex sources API，获取 2yr影响因子和 h_index。
    source_ids: list of OpenAlex source IDs（形如 "S12345678"，去掉 URL 前缀）
    返回 {source_id: {"2yr_if": float, "h_index": int}}
    """
    if not source_ids:
        return {}

    result = {}
    # OpenAlex source filter: pipe-separated IDs, 每次最多50个
    chunk_size = 50
    for i in range(0, len(source_ids), chunk_size):
        chunk = source_ids[i:i + chunk_size]
        filter_str = "id:" + "|".join(chunk)
        params = {
            "filter": filter_str,
            "select": "id,display_name,summary_stats",
            "per-page": chunk_size,
        }
        data = openalex_get("sources", params)
        if data:
            for src in data.get("results", []):
                sid = src.get("id", "").replace("https://openalex.org/", "")
                stats = src.get("summary_stats") or {}
                result[sid] = {
                    "2yr_if": stats.get("2yr_mean_citedness"),
                    "h_index": stats.get("h_index"),
                }
        time.sleep(0.15)

    return result


def enrich_papers_with_journal_stats(papers):
    """
    为论文列表补充期刊分区信息：
    1. 先用内置分区表匹配（精确）
    2. 未匹配则查询 OpenAlex sources API 获取 2yr影响因子，估算分区
    返回 {paper_id: {"jcr_q": str, "cas_tier": str, "2yr_if": float}}
    """
    # 收集需要查询的 source IDs
    source_id_map = {}  # paper_id → source_id
    need_api = []       # source IDs 未在内置表中找到

    paper_journal_map = {}  # paper_id → journal_name（已小写）

    for paper in papers:
        pid = paper.get("id", "")
        primary_loc = paper.get("primary_location") or {}
        source = primary_loc.get("source") or {}
        journal_name = (source.get("display_name") or "").lower().strip()
        source_id = source.get("id", "").replace("https://openalex.org/", "")

        paper_journal_map[pid] = (journal_name, source_id)

        # 检查内置分区表
        if journal_name not in JOURNAL_RANKING_TABLE and source_id:
            if source_id not in need_api:
                need_api.append(source_id)

    # 批量查询未收录期刊的 OpenAlex source stats
    print(f"  [分区] 内置表已覆盖 {len(papers) - len(set(sid for _, sid in paper_journal_map.values() if sid in [s for s in need_api]))}"
          f" 篇，查询 {len(need_api)} 个未收录期刊的影响因子...")
    api_stats = fetch_source_stats(need_api) if need_api else {}

    # 构建结果
    enriched = {}
    for paper in papers:
        pid = paper.get("id", "")
        journal_name, source_id = paper_journal_map.get(pid, ("", ""))

        if journal_name in JOURNAL_RANKING_TABLE:
            jcr_q, cas_tier = JOURNAL_RANKING_TABLE[journal_name]
            impact = api_stats.get(source_id, {}).get("2yr_if") if source_id else None
        else:
            # 用 API 数据估算
            stats = api_stats.get(source_id, {}) if source_id else {}
            impact = stats.get("2yr_if")
            jcr_q, cas_tier = infer_quartile_from_impact(impact)

        enriched[pid] = {
            "jcr_q": jcr_q,
            "cas_tier": cas_tier,
            "2yr_if": round(impact, 2) if impact else "",
        }

    return enriched


def search_openalex(keyword_groups, max_results=100,
                    social_science_filter=True,
                    year_from=None, year_to=None):
    """
    多组 AND 联合检索策略：
    - 以最具体的组（词数最少的组）作为主检索，获取足够多结果
    - 对结果的 title+abstract 做后置过滤：确保包含每个其他组的至少一个同义词
    - 按被引量排序输出
    """
    print(f"\n关键词分组: {keyword_groups}")
    print(f"社会科学过滤: {social_science_filter}  年份范围: {year_from or '不限'}-{year_to or '不限'}")

    # 解析每组的同义词列表
    parsed_groups = []
    for group in keyword_groups:
        synonyms = [s.strip() for s in group.split(" OR ") if s.strip()]
        parsed_groups.append([s.lower() for s in synonyms])

    # 选最具体的组作为主检索（同义词最少 = 最精确）
    primary_idx = min(range(len(parsed_groups)), key=lambda i: len(parsed_groups[i]))
    primary_synonyms = parsed_groups[primary_idx]
    filter_groups = [parsed_groups[i] for i in range(len(parsed_groups)) if i != primary_idx]

    print(f"\n主检索组（索引{primary_idx+1}）: {keyword_groups[primary_idx][:60]}...")

    # 构造 OpenAlex 过滤器
    filter_parts = []
    if social_science_filter:
        filter_parts.append("concepts.id:" + "|".join(SOCIAL_SCIENCE_CONCEPTS))
    if year_from:
        filter_parts.append(f"from_publication_date:{year_from}-01-01")
    if year_to:
        filter_parts.append(f"to_publication_date:{year_to}-12-31")
    filter_str = ",".join(filter_parts) if filter_parts else None

    # 主检索：对每个主组同义词分别检索，合并结果
    primary_map = {}
    total_ref = 0
    for syn in primary_synonyms:
        query = f'"{syn}"' if " " in syn else syn
        params = {
            "search": query,
            "per-page": 200,
            "select": "id,title,authorships,publication_year,primary_location,"
                      "cited_by_count,abstract_inverted_index,doi,open_access,topics",
            "sort": "cited_by_count:desc",
        }
        if filter_str:
            params["filter"] = filter_str

        cursor = "*"
        fetched = 0
        while fetched < 1000:  # 每个同义词最多取1000篇用于过滤
            params["cursor"] = cursor
            data = openalex_get("works", params)
            if not data:
                break
            items = data.get("results", [])
            meta = data.get("meta", {})
            total_ref = max(total_ref, meta.get("count", 0))
            for item in items:
                pid = item.get("id")
                if pid and pid not in primary_map:
                    primary_map[pid] = item
            fetched += len(items)
            next_cursor = meta.get("next_cursor")
            if not next_cursor or len(items) < 200 or fetched >= 1000:
                break
            cursor = next_cursor
            time.sleep(0.12)
        time.sleep(0.15)

    print(f"  主检索合并: {len(primary_map)} 篇")

    if not filter_groups:
        papers = sorted(primary_map.values(),
                        key=lambda x: x.get("cited_by_count", 0), reverse=True)
        return papers[:max_results], len(primary_map)

    # 后置过滤：title+abstract 必须包含每个过滤组的至少一个同义词
    def text_contains_group(paper, group_synonyms):
        title = (paper.get("title") or "").lower()
        abstract = decode_abstract(paper.get("abstract_inverted_index")).lower()
        full_text = title + " " + abstract
        return any(syn in full_text for syn in group_synonyms)

    passed = []
    for pid, paper in primary_map.items():
        if all(text_contains_group(paper, grp) for grp in filter_groups):
            passed.append(paper)

    print(f"  后置过滤后: {len(passed)} 篇（需包含所有 {len(filter_groups)} 个额外组的词）")

    # 若过滤后 < 20 篇，放宽：只要求包含任意一个额外组
    if len(passed) < 20 and len(filter_groups) > 1:
        passed_relaxed = []
        for pid, paper in primary_map.items():
            if any(text_contains_group(paper, grp) for grp in filter_groups):
                passed_relaxed.append(paper)
        print(f"  [自动放宽] 任意一个额外组匹配: {len(passed_relaxed)} 篇")
        passed = passed_relaxed

    passed.sort(key=lambda x: x.get("cited_by_count", 0), reverse=True)
    return passed[:max_results], len(passed)


def format_paper(paper, category="", journal_stats=None):
    """将 OpenAlex paper 对象格式化为输出行"""
    title = paper.get("title") or ""

    # 作者（最多3位）
    authorships = paper.get("authorships") or []
    authors = [a.get("author", {}).get("display_name", "") for a in authorships[:3]]
    if len(authorships) > 3:
        authors.append("et al.")
    author_str = "; ".join(filter(None, authors))

    # 期刊
    primary_loc = paper.get("primary_location") or {}
    source = primary_loc.get("source") or {}
    journal = source.get("display_name") or ""

    # 期刊分区信息
    pid = paper.get("id", "")
    stats = (journal_stats or {}).get(pid, {})
    jcr_q = stats.get("jcr_q", "")
    cas_tier = stats.get("cas_tier", "")
    impact_2yr = stats.get("2yr_if", "")

    # 顶刊标注（分区已知且为Q1则加星）
    is_top = jcr_q.startswith("Q1") or journal.lower() in {
        k for k, v in JOURNAL_RANKING_TABLE.items() if v[0] == "Q1"
    }
    journal_display = f"⭐ {journal}" if is_top else journal

    # 年份、被引
    year = paper.get("publication_year") or ""
    cited = paper.get("cited_by_count") or 0

    # DOI
    doi = paper.get("doi") or ""

    # OA 链接
    oa_info = paper.get("open_access") or {}
    oa_url = oa_info.get("oa_url") or ""

    # 摘要
    abstract = decode_abstract(paper.get("abstract_inverted_index"))

    # 主题标签
    topics = paper.get("topics") or []
    topic_names = [t.get("display_name", "") for t in topics[:3]]
    topics_str = "; ".join(filter(None, topic_names))

    return {
        "文献类别": category,
        "标题": title,
        "作者": author_str,
        "期刊": journal_display,
        "JCR分区": jcr_q,
        "中科院分区": cas_tier,
        "2yr影响因子": impact_2yr,
        "年份": year,
        "被引量": cited,
        "DOI": doi,
        "OA全文链接": oa_url,
        "主题标签": topics_str,
        "摘要": abstract,
    }


def filter_by_quartile(papers, min_quartile):
    """
    按最低分区筛选（Q1最严，Q4最宽）
    min_quartile: "Q1" / "Q2" / "Q3" / "Q4"
    带 * 号的估算值也参与筛选（保守处理：估算Q1*视为Q1可通过）
    """
    order = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}
    min_val = order.get(min_quartile.upper(), 4)

    filtered = []
    for p in papers:
        q = p.get("JCR分区", "").rstrip("*")
        val = order.get(q, 99)
        if val <= min_val:
            filtered.append(p)
    return filtered


def save_excel(all_batches, output_path, topic_label=""):
    """
    all_batches: list of (category_name, color_hex, papers)
    papers 每项包含 JCR分区、中科院分区、2yr影响因子 字段
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "文献汇总"

    headers = ["序号", "文献类别", "标题", "作者", "期刊",
               "JCR分区", "中科院分区", "2yr影响因子",
               "年份", "被引量", "DOI", "OA全文链接", "主题标签", "摘要"]
    ws.append(headers)

    # 表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    seen_titles = set()
    seq = 1

    for category, color, papers in all_batches:
        row_fill = PatternFill(fill_type="solid", fgColor=color)
        for paper in papers:
            title = paper.get("标题", "")
            if not title or title in seen_titles:
                continue
            seen_titles.add(title)

            row = [
                seq,
                category,
                paper.get("标题", ""),
                paper.get("作者", ""),
                paper.get("期刊", ""),
                paper.get("JCR分区", ""),
                paper.get("中科院分区", ""),
                paper.get("2yr影响因子", ""),
                paper.get("年份", ""),
                paper.get("被引量", 0),
                paper.get("DOI", ""),
                paper.get("OA全文链接", ""),
                paper.get("主题标签", ""),
                paper.get("摘要", ""),
            ]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.fill = row_fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            # JCR分区列特殊高亮
            jcr_cell = ws.cell(row=ws.max_row, column=6)
            q_val = str(paper.get("JCR分区", "")).rstrip("*")
            q_colors = {"Q1": "C6EFCE", "Q2": "FFEB9C", "Q3": "FCE4D6", "Q4": "F4CCCC"}
            if q_val in q_colors:
                jcr_cell.fill = PatternFill(fill_type="solid", fgColor=q_colors[q_val])

            seq += 1

    # 列宽
    col_widths = [4, 24, 38, 20, 22, 7, 8, 8, 6, 7, 32, 32, 28, 55]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # 行高
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 65

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # 分类统计 sheet
    ws2 = wb.create_sheet("分类统计")
    ws2.append(["文献类别", "颜色说明", "篇数"])
    color_desc = {"D9E1F2": "蓝-直接相关", "E2EFDA": "绿-背景文献", "FFF2CC": "黄-理论文献", "FCE4D6": "橙-扩展视角"}
    for cat, color, papers in all_batches:
        valid = sum(1 for p in papers if p.get("标题"))
        ws2.append([cat, color_desc.get(color, color), valid])
    ws2.append(["合计（去重后）", "", seq - 1])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 20

    # WoS 检索式 sheet
    ws3 = wb.create_sheet("WoS检索式参考")
    ws3.append(["说明", "本检索式可在 Web of Science 高级检索中使用，限定 SSCI 数据库"])
    ws3.append([])
    ws3.append(["本次检索主题", topic_label])
    ws3.append([])
    ws3.append(["分区说明", "JCR分区/中科院分区：有*号的为根据2yr影响因子估算，非官方数据"])
    ws3.append(["数据来源", "内置分区表 + OpenAlex source API (2yr_mean_citedness)"])

    wb.save(output_path)
    return seq - 1


def main():
    parser = argparse.ArgumentParser(description="OpenAlex 学术文献检索工具（含期刊分区）")
    parser.add_argument("--keywords", action="append", required=True,
                        help="关键词组（每个 --keywords 为一组，组内用 + 分隔同义词，组间取 AND 交集）")
    parser.add_argument("--max-results", type=int, default=100, help="最多返回篇数（默认100）")
    parser.add_argument("--year-from", type=int, help="发表年份下限（如 2015）")
    parser.add_argument("--year-to", type=int, help="发表年份上限（如 2025）")
    parser.add_argument("--no-soc-filter", action="store_true", help="关闭社会科学概念过滤")
    parser.add_argument("--min-quartile", default="", choices=["Q1", "Q2", "Q3", "Q4", ""],
                        help="最低期刊分区要求（如 --min-quartile Q2 则只保留Q1/Q2期刊）")
    parser.add_argument("--no-journal-stats", action="store_true",
                        help="跳过期刊分区查询（加快速度）")
    parser.add_argument("--category", default="直接相关", help="文献类别标签（用于分类汇总）")
    parser.add_argument("--color", default="D9E1F2", help="Excel 行颜色（十六进制，默认蓝色）")
    parser.add_argument("--output-dir", default="~/Downloads", help="输出目录")
    parser.add_argument("--output-file", help="指定输出文件路径（覆盖 --output-dir）")
    parser.add_argument("--topic", default="", help="检索主题描述（用于文件名和 WoS sheet）")
    args = parser.parse_args()

    print("\n===== OpenAlex 学术文献检索工具 =====")

    # 解析关键词组（+ 分隔同义词，各组内 OR，组间 AND 交集）
    keyword_groups = []
    for kw_str in args.keywords:
        synonyms = [k.strip() for k in kw_str.split("+") if k.strip()]
        query = " OR ".join(synonyms)
        keyword_groups.append(query)

    soc_filter = not args.no_soc_filter

    # 执行检索
    papers, count = search_openalex(
        keyword_groups,
        max_results=args.max_results,
        social_science_filter=soc_filter,
        year_from=args.year_from,
        year_to=args.year_to,
    )

    print(f"\n[✓] 检索完成，获取 {len(papers)} 篇")

    # 期刊分区信息
    journal_stats = {}
    if not args.no_journal_stats and papers:
        print("\n[...] 查询期刊分区信息...")
        journal_stats = enrich_papers_with_journal_stats(papers)

    # 格式化
    formatted = [format_paper(p, args.category, journal_stats) for p in papers]

    # 按分区筛选（可选）
    if args.min_quartile:
        before = len(formatted)
        formatted = filter_by_quartile(formatted, args.min_quartile)
        print(f"  [分区筛选] {args.min_quartile} 及以上: {before} → {len(formatted)} 篇")

    # 分区分布统计
    if journal_stats:
        q_counter = Counter()
        for p in formatted:
            q = p.get("JCR分区", "未知").rstrip("*") or "未知"
            q_counter[q] += 1
        dist = "  ".join(f"{q}:{n}" for q, n in sorted(q_counter.items()))
        print(f"  期刊分区分布: {dist}")

    # 输出路径
    if args.output_file:
        output_path = Path(args.output_file).expanduser()
    else:
        today = date.today().strftime("%Y%m%d")
        topic_slug = (args.topic or args.keywords[0])[:20].replace(" ", "_").replace('"', '').replace("/", "_")
        filename = f"外文文献检索_{topic_slug}_{today}.xlsx"
        output_path = Path(args.output_dir).expanduser() / filename

    # 保存 Excel
    batches = [(args.category, args.color, formatted)]
    total_saved = save_excel(batches, output_path, topic_label=args.topic)

    print(f"\n[✓] 已保存: {output_path}（共 {total_saved} 篇）")
    print(f"OUTPUT_FILE:{output_path}")


if __name__ == "__main__":
    main()
