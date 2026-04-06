#!/usr/bin/env python3
"""
Export grounded coding results to a formatted Excel file.

Usage:
    python3 export_coding.py --output output.xlsx --json-file /tmp/data.json
    python3 export_coding.py --output output.xlsx --json '<json_string>'
"""

import argparse
import json
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# Style constants
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="2F5496")
BODY_FONT = Font(name="Microsoft YaHei", size=10)
BOLD_FONT = Font(name="Microsoft YaHei", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LIGHT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
CATEGORY_FILLS = [
    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # green
    PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # orange
    PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),  # blue
    PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # yellow
    PatternFill(start_color="E2D9F3", end_color="E2D9F3", fill_type="solid"),  # purple
    PatternFill(start_color="D6F0EF", end_color="D6F0EF", fill_type="solid"),  # teal
    PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),  # red
    PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid"),  # lavender
    PatternFill(start_color="CFE2CF", end_color="CFE2CF", fill_type="solid"),  # sage
    PatternFill(start_color="FAD7A0", end_color="FAD7A0", fill_type="solid"),  # peach
]


def apply_header_style(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER


def apply_body_style(ws, row, col_count, alt=False):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BODY_FONT
        cell.alignment = WRAP_ALIGNMENT
        cell.border = THIN_BORDER
        if alt:
            cell.fill = ALT_FILL


def auto_width(ws, col, min_width=10, max_width=60):
    max_len = min_width
    for row in ws.iter_rows(min_col=col, max_col=col):
        for cell in row:
            if cell.value:
                lines = str(cell.value).split("\n")
                for line in lines:
                    max_len = max(max_len, len(line) * 1.2)
    ws.column_dimensions[get_column_letter(col)].width = min(max_len, max_width)


def create_sheet1_coding_table(wb, data):
    """Sheet 1: Complete coding table."""
    ws = wb.active
    ws.title = "完整编码表"

    # Title row
    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="扎根理论编码表")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Metadata row
    meta = data.get("metadata", {})
    meta_text = f"研究领域：{meta.get('research_field', '')} | 研究主题：{meta.get('research_topic', '')} | 被访者：{meta.get('interviewee', '')} | 编码日期：{meta.get('coding_date', '')}"
    ws.merge_cells("A2:D2")
    meta_cell = ws.cell(row=2, column=1, value=meta_text)
    meta_cell.font = Font(name="Microsoft YaHei", size=9, color="666666")
    meta_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = ["句群序号", "原文内容", "开放编码", "类属"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    apply_header_style(ws, 4, len(headers))

    # Build category color map
    categories = list(dict.fromkeys(item["category"] for item in data.get("coding_table", [])))
    cat_color_map = {cat: CATEGORY_FILLS[i % len(CATEGORY_FILLS)] for i, cat in enumerate(categories)}

    # Data rows
    for i, item in enumerate(data.get("coding_table", [])):
        row = 5 + i
        ws.cell(row=row, column=1, value=item.get("id", ""))
        ws.cell(row=row, column=2, value=item.get("original_text", ""))
        ws.cell(row=row, column=3, value=item.get("open_code", ""))
        ws.cell(row=row, column=4, value=item.get("category", ""))
        apply_body_style(ws, row, len(headers), alt=(i % 2 == 1))
        # Color the category column by category
        cat = item.get("category", "")
        if cat in cat_color_map:
            ws.cell(row=row, column=4).fill = cat_color_map[cat]

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # Freeze header
    ws.freeze_panes = "A5"

    # Auto filter
    if data.get("coding_table"):
        last_row = 4 + len(data["coding_table"])
        ws.auto_filter.ref = f"A4:D{last_row}"


def create_sheet2_category_summary(wb, data):
    """Sheet 2: Category summary."""
    ws = wb.create_sheet("编码-类属汇总")

    ws.merge_cells("A1:C1")
    ws.cell(row=1, column=1, value="编码-类属汇总表").font = TITLE_FONT

    headers = ["类属", "包含的开放编码", "编码数量"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    apply_header_style(ws, 3, len(headers))

    for i, cat in enumerate(data.get("categories", [])):
        row = 4 + i
        ws.cell(row=row, column=1, value=cat.get("name", ""))
        codes = cat.get("codes", [])
        ws.cell(row=row, column=2, value="、".join(codes))
        ws.cell(row=row, column=3, value=cat.get("code_count", len(codes)))
        apply_body_style(ws, row, len(headers), alt=(i % 2 == 1))
        ws.cell(row=row, column=1).fill = CATEGORY_FILLS[i % len(CATEGORY_FILLS)]
        ws.cell(row=row, column=1).font = BOLD_FONT

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 12

    ws.freeze_panes = "A4"


def create_sheet3_properties(wb, data):
    """Sheet 3: Properties and dimensions."""
    ws = wb.create_sheet("属性与维度")

    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1, value="类属属性与维度分析").font = TITLE_FONT

    headers = ["类属", "属性1", "属性1维度", "属性2", "属性2维度", "四象限分类"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    apply_header_style(ws, 3, len(headers))

    for i, cat in enumerate(data.get("categories", [])):
        row = 4 + i
        ws.cell(row=row, column=1, value=cat.get("name", ""))

        p1 = cat.get("property1", {})
        ws.cell(row=row, column=2, value=p1.get("name", ""))
        ws.cell(row=row, column=3, value=f"{p1.get('dimension_a', '')} ←→ {p1.get('dimension_b', '')}")

        p2 = cat.get("property2", {})
        ws.cell(row=row, column=4, value=p2.get("name", ""))
        ws.cell(row=row, column=5, value=f"{p2.get('dimension_a', '')} ←→ {p2.get('dimension_b', '')}")

        # Quadrants
        quads = cat.get("quadrants", {})
        quad_text_parts = []
        for key in ["type_I", "type_II", "type_III", "type_IV"]:
            q = quads.get(key, {})
            label = {"type_I": "I", "type_II": "II", "type_III": "III", "type_IV": "IV"}.get(key, key)
            quad_text_parts.append(f"类型{label}：{q.get('name', '')}（{q.get('description', '')}）")
        ws.cell(row=row, column=6, value="\n".join(quad_text_parts))

        apply_body_style(ws, row, len(headers), alt=(i % 2 == 1))
        ws.cell(row=row, column=1).fill = CATEGORY_FILLS[i % len(CATEGORY_FILLS)]
        ws.cell(row=row, column=1).font = BOLD_FONT

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 50

    ws.freeze_panes = "A4"


def create_sheet4_statistics(wb, data):
    """Sheet 4: Coding statistics."""
    ws = wb.create_sheet("编码统计")

    ws.merge_cells("A1:B1")
    ws.cell(row=1, column=1, value="编码统计摘要").font = TITLE_FONT

    stats = data.get("statistics", {})

    # Summary stats
    summary_items = [
        ("总句群数", stats.get("total_segments", 0)),
        ("开放编码数（去重）", stats.get("unique_codes", 0)),
        ("类属总数", stats.get("total_categories", 0)),
    ]

    row = 3
    for label, value in summary_items:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2, value=value).font = BODY_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # Category distribution
    row += 1
    ws.cell(row=row, column=1, value="类属分布").font = SUBTITLE_FONT
    row += 1
    dist_headers = ["类属", "编码数量", "占比"]
    for col, h in enumerate(dist_headers, 1):
        ws.cell(row=row, column=col, value=h)
    apply_header_style(ws, row, len(dist_headers))
    row += 1

    for i, item in enumerate(stats.get("category_distribution", [])):
        ws.cell(row=row, column=1, value=item.get("category", ""))
        ws.cell(row=row, column=2, value=item.get("code_count", 0))
        ws.cell(row=row, column=3, value=item.get("percentage", ""))
        apply_body_style(ws, row, len(dist_headers), alt=(i % 2 == 1))
        row += 1

    # Top codes
    row += 1
    ws.cell(row=row, column=1, value="高频编码 Top 5").font = SUBTITLE_FONT
    row += 1
    top_headers = ["编码", "出现频次"]
    for col, h in enumerate(top_headers, 1):
        ws.cell(row=row, column=col, value=h)
    apply_header_style(ws, row, len(top_headers))
    row += 1

    for i, item in enumerate(stats.get("top_codes", [])):
        ws.cell(row=row, column=1, value=item.get("code", ""))
        ws.cell(row=row, column=2, value=item.get("frequency", 0))
        apply_body_style(ws, row, len(top_headers), alt=(i % 2 == 1))
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12


def main():
    parser = argparse.ArgumentParser(description="Export grounded coding results to Excel")
    parser.add_argument("--output", required=True, help="Output Excel file path")
    parser.add_argument("--json", help="JSON data string")
    parser.add_argument("--json-file", help="Path to JSON data file")
    args = parser.parse_args()

    if args.json_file:
        with open(args.json_file, "r", encoding="utf-8") as f:
            data = json.load(f)
    elif args.json:
        data = json.loads(args.json)
    else:
        print("Error: provide --json or --json-file")
        sys.exit(1)

    wb = Workbook()
    create_sheet1_coding_table(wb, data)
    create_sheet2_category_summary(wb, data)
    create_sheet3_properties(wb, data)
    create_sheet4_statistics(wb, data)

    wb.save(args.output)
    print(f"Excel file saved to: {args.output}")


if __name__ == "__main__":
    main()
